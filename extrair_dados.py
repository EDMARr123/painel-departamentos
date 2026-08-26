r"""
Extrai os dados do painel "Departamentos" a partir de RESULTADO.xlsx
(aba "MES ATUAL") e salva um dados.json pronto pro gerador de HTML
consumir.

Layout da aba: blocos por supervisor (linha com nome do supervisor na
col C, ex: "EDMAR"), seguidos de uma linha de cabeçalho ("RCA" / "VENDEDOR")
e 1 linha por RCA até a linha "TOTAL" que fecha o bloco.

Por RCA (só as 10 categorias de produto — Mix Médio e SKU ficam de fora):
- C = código, D = nome completo (formato "NOME - ROTA...", às vezes só "NOME")
- N = BACON realizado             | meta fixa na linha 5, col N
- P = BOVINO realizado            | meta fixa na linha 5, col P
- R = BATATA realizado            | meta fixa na linha 5, col R
- T = SUINO realizado             | meta fixa na linha 5, col T
- V = CALABRESA realizado         | meta fixa na linha 5, col V
- X = PÃES realizado              | meta fixa na linha 5, col X
- Z = FRESCAIS realizado          | meta fixa na linha 5, col Z
- AB = SABORIZADAS realizado      | meta não tem célula numérica própria —
- AD = LACTEOS realizado          |   vem do texto "MINIMO N" no cabeçalho
- AM = THERMO realizado           |   de cada bloco (col AB/AD/AM da linha do supervisor)

A linha 5 (mínimos "oficiais") é global — vale pra todos os blocos/supervisores,
mesmo que o texto do cabeçalho de algum bloco mostre um número desatualizado
(ex: bloco do RICHARD mostra "MINIMO 2,15" no texto pra Mix Médio, que nem
entra mais aqui — mas confirma que o texto do cabeçalho pode ficar desatualizado).

"Bateu a meta" (status do card) = realizado >= meta em TODAS as 10 categorias
de produto.
"""

import json
import os
import re

import openpyxl

CAMINHO_RESULTADO = r"C:\Users\edmar\Desktop\ACOMPANHA RESULTADO\RESULTADO.xlsx"
CAMINHO_MELHORIA_SALARIAL = r"c:\AutomacaoMaxGestao\melhoria_salarial\dados.json"
CAMINHO_PAINEL_PILARES = r"c:\AutomacaoMaxGestao\painel_pilares\dados.json"

PASTA_BASE = os.path.dirname(os.path.abspath(__file__))
CAMINHO_SAIDA = os.path.join(PASTA_BASE, "dados.json")

# Nome do supervisor na planilha -> nome de exibição/agrupamento no painel.
NORMALIZAR_SUPERVISOR = {
    "EDMAR": "EDMAR",
    "LEANDRO FREITAS": "LEANDRO",
    "FLAVIANE": "FLAVIANE",
    "IDEGLAN": "IDEGLAN",
    "RICARDO": "RICARDO",
    "SUP RICHARD": "RICHARD",
    "RODRIGO": "RODRIGO",
}

# (chave, rótulo de exibição, coluna do realizado)
# Mix Médio (J) e SKU (L) ficam de fora — o painel mostra só as categorias
# de produto.
CATEGORIAS = [
    ("bacon", "Bacon", 14),           # N
    ("bovino", "Bovino", 16),         # P
    ("batata", "Batata", 18),         # R
    ("suino", "Suíno", 20),           # T
    ("calabresa", "Calabresa", 22),   # V
    ("paes", "Pães", 24),             # X
    ("frescais", "Frescais", 26),     # Z
    ("lacteos", "Lácteos", 30),       # AD
    ("thermo", "Thermo", 39),         # AM
]
# "saborizadas" (col AB) retirada do painel a pedido do Edmar (25/08) —
# fica de fora da contagem "bateu a meta" também.

# Até 19/08 só saborizadas/AB, lacteos/AD e thermo/AM tinham o texto
# "MINIMO N" no cabeçalho de cada bloco — as outras 7 categorias usavam um
# valor numérico fixo da linha 5 (mesmo pra todo mundo). Confirmado em 21/08
# que agora TODO bloco de supervisor tem seu próprio "MINIMO N" pras 10
# categorias (e alguns minimos mudaram, ex: Bovino é 15 e não mais 10) —
# a linha 5 ficou obsoleta, a meta de cada categoria sempre vem do texto do
# cabeçalho do bloco correspondente.


def _num(v):
    return v if isinstance(v, (int, float)) else 0


def _parse_minimo(texto):
    """'MINIMO 15' -> 15.0 · 'MINIMO 2,50' -> 2.5 · None/outro -> 0.0"""
    if not isinstance(texto, str):
        return 0.0
    m = re.search(r"[\d.,]+", texto)
    if not m:
        return 0.0
    return float(m.group(0).replace(",", "."))


def _nome_e_rota(nome_completo):
    """'FABIO L. - GYN RT 21 - SEG 01' -> ('FABIO L.', 'GYN RT 21 - SEG 01')
    Nomes sem ' - ' (com espaços dos dois lados) ficam só com o nome, sem rota."""
    partes = [p.strip() for p in nome_completo.split(" - ")]
    nome = partes[0].strip()
    rota = " - ".join(p for p in partes[1:] if p).strip()
    return nome, rota


def _ler_media_pedidos_atual():
    """Cross-referencia com o painel Performance (melhoria_salarial) pra
    pegar a média de pedidos/dia do mês ATUAL de cada RCA (total de
    pedidos do mês / dias úteis) — precisa rodar depois do
    melhoria_salarial/extrair_dados.py pra pegar o dado mais recente."""
    if not os.path.exists(CAMINHO_MELHORIA_SALARIAL):
        return {}
    with open(CAMINHO_MELHORIA_SALARIAL, "r", encoding="utf-8") as f:
        dados = json.load(f)
    dias_uteis = dados["constantes"]["dias_uteis"] or 1
    return {
        str(r["codigo"]): round(r["total_pedidos"] / dias_uteis)
        for r in dados["rcas"]
    }


def _ler_posit_atual():
    """Cross-referencia com o Painel 4 Pilares pra pegar a 'Média de
    pedidos' (coluna BD da SOMA NAO SALVA ENCIMA.xlsx) de cada RCA — é uma
    métrica diferente da 'média pedidos mês atual' (essa vem do total de
    pedidos ÷ dias úteis calculado no Performance)."""
    if not os.path.exists(CAMINHO_PAINEL_PILARES):
        return {}
    with open(CAMINHO_PAINEL_PILARES, "r", encoding="utf-8") as f:
        dados = json.load(f)
    return {str(r["codigo"]): r["media_pedidos"] for r in dados}


def extrair():
    wb = openpyxl.load_workbook(CAMINHO_RESULTADO, data_only=True)
    ws = wb["MES ATUAL"]

    def val(row, col):
        return _num(ws.cell(row=row, column=col).value)

    media_pedidos_atual = _ler_media_pedidos_atual()
    posit_atual = _ler_posit_atual()

    rcas = []
    supervisor_atual = None
    metas_bloco = {}

    for r in range(1, ws.max_row + 1):
        c3 = ws.cell(row=r, column=3).value
        c4 = ws.cell(row=r, column=4).value

        if isinstance(c3, str) and c4 is None and c3 not in ("TOTAL",):
            # Linha de cabeçalho de bloco (nome do supervisor)
            supervisor_atual = NORMALIZAR_SUPERVISOR.get(c3.strip(), c3.strip())
            metas_bloco = {chave: _parse_minimo(ws.cell(row=r, column=col).value) for chave, _, col in CATEGORIAS}
            continue

        if not isinstance(c3, int) or not c4 or supervisor_atual is None:
            continue

        nome_rca, rota = _nome_e_rota(str(c4))

        categorias_dados = {}
        atingidas = 0
        for chave, label, col in CATEGORIAS:
            meta = metas_bloco.get(chave, 0)
            real = val(r, col)
            bateu_categoria = real >= meta if meta else False
            if bateu_categoria:
                atingidas += 1
            categorias_dados[chave] = {"label": label, "meta": meta, "real": real, "bateu": bateu_categoria}

        rcas.append({
            "codigo": str(c3),
            "nome": nome_rca,
            "rota": rota,
            "supervisor": supervisor_atual,
            "categorias": categorias_dados,
            "categorias_atingidas": atingidas,
            "total_categorias": len(CATEGORIAS),
            "bateu": atingidas == len(CATEGORIAS),
            "media_pedidos_atual": media_pedidos_atual.get(str(c3), 0),
            "posit_atual": posit_atual.get(str(c3), 0),
        })

    return rcas


if __name__ == "__main__":
    dados = extrair()
    with open(CAMINHO_SAIDA, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)
    print(f"{len(dados)} RCAs extraídos. Salvo em: {CAMINHO_SAIDA}")
